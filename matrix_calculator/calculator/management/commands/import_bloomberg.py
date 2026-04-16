"""
Management command: import_bloomberg

Reads the Bloomberg_Monthly sheet from Matrix_Calcs_3.00.xlsm and
bulk-inserts rows into the BloombergRate table. Idempotent — skips dates
that already exist.

Usage:
    python manage.py import_bloomberg
    python manage.py import_bloomberg --file "C:/path/to/Matrix_Calcs_3.00.xlsm"
"""

import openpyxl
from datetime import date
from django.core.management.base import BaseCommand, CommandError

from calculator.models import BloombergRate, TERM_BUCKETS, TERM_TO_FIELD

# Default path to the workbook
DEFAULT_PATH = r"C:\Users\mchad\Downloads\Matrix_Calcs_3.00.xlsm"

# Column indices in Bloomberg_Monthly sheet (0-based after skipping col A index)
# Row 2: term labels  Row 3: instrument names  Row 4+: data
# Col B = date (bloomberg date), Col C = end-of-month date, Cols D-R = rates
COL_EOM_DATE = 2   # column C (0-based index in row tuple)
COL_RATE_START = 3  # column D onwards


class Command(BaseCommand):
    help = 'Import Bloomberg CDS historical rates from Matrix_Calcs_3.00.xlsm'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file', default=DEFAULT_PATH,
            help='Path to Matrix_Calcs_3.00.xlsm',
        )

    def handle(self, *args, **options):
        path = options['file']
        self.stdout.write(f'Opening {path} ...')

        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except FileNotFoundError:
            raise CommandError(f'File not found: {path}')

        if 'Bloomberg_Monthly' not in wb.sheetnames:
            raise CommandError('Bloomberg_Monthly sheet not found in workbook')

        ws = wb['Bloomberg_Monthly']
        rows = list(ws.iter_rows(min_row=1, values_only=True))

        if len(rows) < 4:
            raise CommandError('Bloomberg_Monthly sheet has too few rows')

        # Row index 1 (0-based) = term labels row (1M, 2M, ... 180M)
        term_row = rows[1]  # row 2 in Excel
        # Build column → term_months mapping from the header
        col_to_term = {}
        for col_idx, cell in enumerate(term_row):
            if cell is None:
                continue
            label = str(cell).strip().upper().replace(' ', '')
            term_map = {
                '1M': 1, '2M': 2, '3M': 3, '6M': 6, '12M': 12,
                '24M': 24, '36M': 36, '48M': 48, '60M': 60,
                '72M': 72, '84M': 84, '96M': 96, '108M': 108,
                '120M': 120, '180M': 180,
            }
            if label in term_map:
                col_to_term[col_idx] = term_map[label]

        self.stdout.write(f'Found term columns: {sorted(col_to_term.values())}')

        existing_dates = set(BloombergRate.objects.values_list('date', flat=True))
        to_create = []
        skipped = 0
        errors = 0

        # Data starts at row index 3 (row 4 in Excel)
        for row in rows[3:]:
            if not row or row[COL_EOM_DATE] is None:
                continue

            # Parse end-of-month date
            raw_date = row[COL_EOM_DATE]
            if hasattr(raw_date, 'date'):
                eom = raw_date.date()
            elif isinstance(raw_date, date):
                eom = raw_date
            else:
                try:
                    from datetime import datetime
                    eom = datetime.strptime(str(raw_date)[:10], '%Y-%m-%d').date()
                except Exception:
                    errors += 1
                    continue

            if eom in existing_dates:
                skipped += 1
                continue

            obj = BloombergRate(date=eom)
            for col_idx, term in col_to_term.items():
                if col_idx < len(row) and row[col_idx] is not None:
                    try:
                        val = float(row[col_idx])
                        field = TERM_TO_FIELD[term]
                        setattr(obj, field, val)
                    except (ValueError, TypeError, KeyError):
                        pass

            to_create.append(obj)
            existing_dates.add(eom)

        if to_create:
            BloombergRate.objects.bulk_create(to_create)
            self.stdout.write(self.style.SUCCESS(
                f'Imported {len(to_create)} Bloomberg rate rows.'
            ))
        else:
            self.stdout.write('No new rows to import.')

        if skipped:
            self.stdout.write(f'Skipped {skipped} already-existing dates.')
        if errors:
            self.stdout.write(self.style.WARNING(f'{errors} rows had unparseable dates.'))
